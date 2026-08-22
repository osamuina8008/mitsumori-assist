#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
見積算定書ジェネレーター
  炉組みアシスト（HTML）が書き出した JSON を、既存テンプレ 0-2.見積内訳表「3」.xlsx に
  流し込み、上司提出用の算定書 .xlsx を生成する。
  各品番（＝図面）ごとに条件が異なってよい。数式（単価/ヶ・合計単価・時間単価・評価等）は
  テンプレのものをそのまま生かすため、入力セルだけを書き込む。

使い方:
  python3 算定書生成.py <入力.json> [出力先フォルダ]
  引数なしの場合は ~/Downloads の最新 算定書_*.json を使う（.command 用）。
"""
import sys, os, json, glob, datetime, shutil
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(HERE, "0-2.見積内訳表「3」.xlsx")

# 付帯工程 → テンプレ列
# 付帯工程 → テンプレ列（工数は副行 r+1 に記入。品番行 T5..Z5 は
#  =工数×加工単価(row3)÷ロット数 の数式なので触らない）。運送費AAは重量×係数で自動。
FUKUTAI = [("防炭","T"),("段取り","U"),("高温プレス焼戻","V"),("サブゼロ","W"),
           ("曲矯正","X"),("SB","Y"),("検査","Z")]

def find_latest_json():
    dl = os.path.expanduser("~/Downloads")
    cands = glob.glob(os.path.join(dl, "算定書_*.json")) + glob.glob(os.path.join(dl, "算定書*.json"))
    if not cands:
        raise SystemExit("~/Downloads に 算定書_*.json が見つかりません。ツールから書き出してください。")
    return max(cands, key=os.path.getmtime)

def num(v, default=0):
    try:
        if v in (None, ""): return default
        return float(v)
    except (TypeError, ValueError):
        return default

def main():
    args = [a for a in sys.argv[1:]]
    src = args[0] if args else find_latest_json()
    outdir = args[1] if len(args) > 1 else os.path.dirname(os.path.abspath(src))

    with open(src, encoding="utf-8") as f:
        data = json.load(f)
    meta  = data.get("meta", {})
    items = data.get("items", [])
    if not items:
        raise SystemExit("items が空です。")

    if not os.path.exists(TEMPLATE):
        raise SystemExit(f"テンプレが見つかりません: {TEMPLATE}")

    wb = openpyxl.load_workbook(TEMPLATE)
    # 品番数に応じてシート選択（5件まで=内訳原紙 / それ以上=13列）
    sheet = "内訳原紙" if len(items) <= 5 else "内訳原紙 (13列）"
    if sheet not in wb.sheetnames:
        sheet = wb.sheetnames[0]
    ws = wb[sheet]
    cap = 5 if sheet == "内訳原紙" else 13
    # 右端に備考列（AI）を新設（テンプレの数式・体裁に影響しない）
    if not ws["AI4"].value:
        ws["AI4"] = "備考"
    if len(items) > cap:
        print(f"⚠ {len(items)}件は {sheet} の上限 {cap} を超過。先頭 {cap} 件のみ記入します。")
        items = items[:cap]

    # ヘッダ（1品番目の行に得意先／日付、各行に営業担当）
    date_str = meta.get("date") or datetime.date.today().isoformat()
    try:
        d_obj = datetime.datetime.strptime(date_str[:10], "%Y-%m-%d").date()
    except ValueError:
        d_obj = datetime.date.today()

    tanto = meta.get("営業担当", "")
    kokyaku = meta.get("得意先", "")

    # テンプレの記入例データを全スロットから消す（入力セルのみ。数式・ラベルは残す）
    ITEM_INPUTS = ["A","B","C","E","F","G","H","I","J","M","N","Q","R","AI"]
    SUB_INPUTS  = ["J","T","U","V","W","X","Y","Z"]
    for j in range(cap):
        rr = 5 + 2*j; ss = rr + 1
        for c in ITEM_INPUTS:
            ws[f"{c}{rr}"] = None
        for c in SUB_INPUTS:
            ws[f"{c}{ss}"] = None

    for i, it in enumerate(items):
        r = 5 + 2*i          # 品番行
        sub = r + 1          # 副行（浸炭B/P）
        ws[f"A{r}"] = d_obj
        ws[f"A{r}"].number_format = "yyyy/mm/dd"
        if tanto:   ws[f"B{r}"] = tanto
        if i == 0 and kokyaku:
            ws[f"C{r}"] = kokyaku
        ws[f"E{r}"] = it.get("品番", "")
        ws[f"F{r}"] = it.get("品名", "")
        ws[f"G{r}"] = it.get("処理名", "")
        ws[f"H{r}"] = num(it.get("ロット数量Max"))
        ws[f"I{r}"] = num(it.get("受注量月"))
        ws[f"J{r}"] = num(it.get("単重量"))
        ws[f"M{r}"] = num(it.get("混載リスク率"), 1)
        ws[f"N{r}"] = num(it.get("合計単価"))
        ws[f"Q{r}"] = num(it.get("処理時間"))
        ws[f"R{r}"] = it.get("処理条件", "")
        ws[f"AI{r}"] = it.get("備考", "")
        # 炉種 B/P → 副行 J
        furnace = str(it.get("炉", "B")).upper()
        ws[f"J{sub}"] = "浸炭P" if furnace == "P" else "浸炭B"
        # 付帯工数（副行に記入 → 品番行の数式が単価換算）
        fk = it.get("付帯", {}) or {}
        for key, col in FUKUTAI:
            v = num(fk.get(key))
            if v:
                ws[f"{col}{sub}"] = v
        # 運送費を明示指定したいとき（テンプレ既定は 単重量×13.3 の自動）
        soryo = it.get("運送費", None)
        if soryo not in (None, ""):
            ws[f"AA{r}"] = num(soryo)

    os.makedirs(outdir, exist_ok=True)
    safe_k = (kokyaku or "見積").replace("/", "_").replace("\\", "_")
    d10 = date_str[:10].replace("-", "")
    out = os.path.join(outdir, f"{d10}_{safe_k}_見積算定書.xlsx")
    n = 1
    base = out
    while os.path.exists(out):
        n += 1
        out = base[:-5] + f"_{n}.xlsx"
    wb.save(out)
    print(f"✅ 生成しました（{len(items)}件 / シート「{sheet}」）:\n{out}")
    return out

if __name__ == "__main__":
    out = main()
    # Mac: 生成後に開く
    if sys.platform == "darwin":
        import subprocess
        subprocess.run(["open", out])
